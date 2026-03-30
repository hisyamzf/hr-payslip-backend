# backend/app/utils/excel_parser.py
"""
Real Excel parser using openpyxl
Supports flexible column mapping and error handling
"""
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

logger = logging.getLogger(__name__)

class ExcelParsingError(Exception):
    """Custom exception untuk Excel parsing errors"""
    pass


class ExcelParser:
    """
    Real Excel parser dengan dukungan untuk:
    - Flexible column mapping (fixed columns, earnings, deductions)
    - Edge case handling (empty cells, string numbers, trimming)
    - Row-level error tracking
    - Support untuk local file dan bytes input
    """
    
    NUMERIC_TYPES = (int, float, Decimal)
    
    def __init__(self, max_rows: int = None, skip_empty_rows: bool = True, start_row: int = 2):
       
        """
        Args:
            max_rows: Limit jumlah rows yang diparse (None = unlimited)
            skip_empty_rows: Skip baris kosong (default: True)
            start_row: Row mulai parse (default: 2)
        """
        self.max_rows = max_rows
        self.skip_empty_rows = skip_empty_rows
        self.start_row = start_row  # <-- row mulai parse, default 2
    
    def parse_file(self, file_path: str, column_mapping: Dict[str, Any], start_row: int = None):
        start_row = start_row or column_mapping.get('start_row', 2)
        wb = load_workbook(file_path, data_only=True)
        
        # Fixed columns sama untuk semua sheet
        fixed_columns = column_mapping.get('fixed_columns', {})
        employee_key = 'employee_number'  # bisa disesuaikan dengan fixed_columns

        rows_dict = {}  # key = employee_number

        # --- Parsing earnings sheet ---
        earnings_mapping = column_mapping.get('earnings', {})
        if earnings_mapping:
            sheet_name = earnings_mapping.get('sheet_name')
            ws = wb[sheet_name]
            for excel_row_idx, excel_row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
                row_values = [cell.value for cell in excel_row]
                emp_id = self._get_cell_value(row_values, self._column_letter_to_index(fixed_columns[employee_key]))
                if emp_id is None:
                    continue  # skip row tanpa employee_number
                
                # Fixed columns
                parsed_row = {k: self._get_cell_value(row_values, self._column_letter_to_index(v)) for k, v in fixed_columns.items()}
                
                # Earnings
                earnings = {}
                for col in earnings_mapping.get('columns', []):
                    idx = self._column_letter_to_index(col['column'])
                    earnings[col['key']] = self._to_numeric(row_values[idx])
                
                parsed_row['earnings'] = earnings
                parsed_row['deductions'] = {}  # placeholder nanti di-update
                parsed_row['row_number'] = excel_row_idx
                rows_dict[emp_id] = parsed_row

        # --- Parsing deductions sheet ---
        deductions_mapping = column_mapping.get('deductions', {})
        if deductions_mapping:
            sheet_name = deductions_mapping.get('sheet_name')
            ws = wb[sheet_name]
            for excel_row_idx, excel_row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
                row_values = [cell.value for cell in excel_row]
                emp_id = self._get_cell_value(row_values, self._column_letter_to_index(fixed_columns[employee_key]))
                if emp_id is None:
                    continue
                
                if emp_id in rows_dict:
                    deductions = {}
                    for col in deductions_mapping.get('columns', []):
                        col_idx = self._column_letter_to_index(col['column'])
                        deductions[col['key']] = self._to_numeric(row_values[col_idx])
                    rows_dict[emp_id]['deductions'] = deductions
                else:
                    # jika employee di deductions tidak ada di earnings
                    print(f"⚠️ Employee {emp_id} di sheet {sheet_name} tidak ada di earnings, skip")

        # Return list rows
        return list(rows_dict.values())
    
    def parse_bytes(
        self,
        file_content: bytes,
        column_mapping: Dict[str, Any],
        filename: str = "upload.xlsx"
    ) -> List[Dict[str, Any]]:
        """
        Parse Excel dari bytes (e.g., dari Supabase Storage upload)
        
        Args:
            file_content: Raw file bytes
            column_mapping: Column mapping config
            filename: Nama file (untuk logging)
        
        Returns:
            Parsed rows (same format seperti parse_file)
        """
        from io import BytesIO
        
        try:
            from openpyxl import load_workbook
            
            logger.info(f"📖 Parsing Excel from bytes: {filename}")
            
            # Load dari bytes
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
            
            if ws is None:
                raise ExcelParsingError("Tidak ada worksheet aktif dalam file Excel")
            
            # Parse rows
            rows = []
            row_count = 0
            
            for excel_row_idx, excel_row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                row_count += 1
                
                if self.max_rows and row_count > self.max_rows:
                    logger.info(f"⏸️  Reached max_rows limit: {self.max_rows}")
                    break
                
                row_values = [cell.value for cell in excel_row]
                
                if self.skip_empty_rows and all(v is None or str(v).strip() == "" for v in row_values):
                    continue
                
                try:
                    parsed_row = self._parse_row(row_values, excel_row_idx, column_mapping)
                    rows.append(parsed_row)
                except Exception as e:
                    logger.warning(f"⚠️  Error parsing row {excel_row_idx}: {str(e)}")
                    continue
            
            logger.info(f"✅ Berhasil parse {len(rows)} rows dari bytes")
            return rows
        
        except Exception as e:
            logger.error(f"❌ Error parsing bytes: {str(e)}")
            raise ExcelParsingError(f"Error parsing bytes: {str(e)}")
    
    def _parse_row(
        self, 
        row_values: List[Any], 
        row_number: int, 
        column_mapping: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Parse single row berdasarkan column mapping
        
        Returns:
            {
                'row_number': 2,
                'employee_number': 'EMP-001',
                'full_name': 'John Doe',
                'earnings': {'salary': 5000000.0},
                'deductions': {'tax': 500000.0}
            }
        """
        parsed = {
            'row_number': row_number,
        }
        
        # Extract fixed columns (employee_number, full_name, etc)
        fixed_columns = column_mapping.get('fixed_columns', {})
        for col_name, col_letter in fixed_columns.items():
            col_idx = self._column_letter_to_index(col_letter)
            value = self._get_cell_value(row_values, col_idx)
            
            # For fixed columns like employee_number, full_name - trim string
            if isinstance(value, str):
                value = value.strip()
            
            parsed[col_name] = value
        
        # Extract earnings (dynamic columns)
        earnings = {}
        earnings_mapping = column_mapping.get('earnings', [])
        for col_info in earnings_mapping:
            col_letter = col_info.get('column')
            col_key = col_info.get('key')
            
            if not col_letter or not col_key:
                continue
            
            col_idx = self._column_letter_to_index(col_letter)
            value = self._get_cell_value(row_values, col_idx)
            
            # Convert to numeric, treat None/empty as 0
            numeric_value = self._to_numeric(value)
            earnings[col_key] = numeric_value
        
        parsed['earnings'] = earnings
        
        # Extract deductions (dynamic columns)
        deductions = {}
        deductions_mapping = column_mapping.get('deductions', [])
        for col_info in deductions_mapping:
            col_letter = col_info.get('column')
            col_key = col_info.get('key')
            
            if not col_letter or not col_key:
                continue
            
            col_idx = self._column_letter_to_index(col_letter)
            value = self._get_cell_value(row_values, col_idx)
            
            # Convert to numeric, treat None/empty as 0
            numeric_value = self._to_numeric(value)
            deductions[col_key] = numeric_value
        
        parsed['deductions'] = deductions
        
        return parsed
    
    def _column_letter_to_index(self, col_letter: str) -> int:
        """
        Convert column letter (A, B, C, ..., Z, AA, AB, etc) ke 0-based index
        
        Examples:
            'A' -> 0
            'B' -> 1
            'Z' -> 25
            'AA' -> 26
            'AB' -> 27
        """
        col_letter = col_letter.upper()
        index = 0
        for i, char in enumerate(reversed(col_letter)):
            index += (ord(char) - ord('A') + 1) * (26 ** i)
        return index - 1
    
    def _get_cell_value(self, row_values: List[Any], col_idx: int) -> Any:
        """
        Safety get cell value dengan bounds checking
        
        Returns:
            Cell value jika exist, None jika out of bounds atau None
        """
        if 0 <= col_idx < len(row_values):
            return row_values[col_idx]
        return None
    
    def _to_numeric(self, value: Any, default: float = 0.0) -> float:
        """
        Convert value ke numeric (float)
        
        Handles:
        - Already numeric (int, float, Decimal) -> return as float
        - String number ("1500000", "1.5") -> parse to float
        - String with comma ("1.500.000") -> remove dots and parse
        - None / empty string -> return default (0.0)
        - Invalid string -> return default dengan warning
        """
        # Return None dan valid numeric values
        if value is None:
            return default
        
        # Already numeric type
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value) if value else default
        
        # String value
        if isinstance(value, str):
            value = value.strip()
            
            if not value:  # Empty string
                return default
            
            try:
                # Handle Indonesian format: "1.500.000" (dots as thousands separator)
                # First check if it uses comma as decimal
                if ',' in value and '.' in value:
                    # "1.500.000,50" -> remove dots from thousands, replace comma with dot
                    value = value.replace('.', '').replace(',', '.')
                elif ',' in value:
                    # "1500,50" or "1,50" - replace comma dengan dot
                    value = value.replace(',', '.')
                
                # Try to convert to float
                return float(value)
            
            except ValueError:
                logger.warning(f"⚠️  Cannot convert '{value}' to numeric, using default {default}")
                return default
        
        # Other types - return default
        return default


    def get_preview(self, file_path: str, max_rows: int = 5, sheet_name: str = None):
        """
        Get preview of Excel file (headers + sample rows)
        
        Args:
            file_path: Path to Excel file
            max_rows: Maximum number of rows to return (default: 5)
            sheet_name: Specific sheet to preview (optional, uses active sheet if not specified)
            
        Returns:
            Tuple of (headers, sample_rows)
        """
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
        else:
            ws = wb.active
        
        headers = []
        sample_rows = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows + 1, values_only=True)):
            if row_idx == 0:
                headers = [str(cell) if cell else "" for cell in row]
            else:
                sample_rows.append([str(cell) if cell else "" for cell in row])
        
        return headers, sample_rows
    
    def get_all_sheets(self, file_path: str) -> List[str]:
        """
        Get all sheet names from Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        wb = load_workbook(file_path, data_only=True)
        return wb.sheetnames
    
    def auto_detect_columns(self, file_path: str) -> dict:
        """
        Auto-detect column mapping from Excel headers
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Auto-detected column mapping dict
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).lower().strip() if cell else "" for cell in row]
        
        mapping = {
            'fixed_columns': {},
            'earnings': [],
            'deductions': [],
        }
        
        # Common column name patterns
        fixed_patterns = {
            'employee_number': ['employee_number', 'emp_no', 'emp number', 'no. karyawan', 'no karyawan', 'nomor karyawan', 'nik', 'nopeg', 'no emp', 'no.pegawai'],
            'full_name': ['full_name', 'name', 'nama', 'nama lengkap', 'employee name', 'nama karyawan'],
            'department': ['department', 'dept', 'departemen', 'bagian'],
            'position': ['position', 'jabatan', 'title', 'job title', 'posisi'],
        }
        
        # Earning column patterns
        earning_patterns = {
            'salary': ['salary', 'gaji', 'gaji pokok', 'basic salary', 'basic_salary'],
            'allowance': ['allowance', 'tunjangan', 'tunjangan pokok', 'transport', 'meal', 'bonus'],
            'overtime': ['overtime', 'lembur', 'lemburan'],
            'bonus': ['bonus', 'thr', 'tunjangan hari raya'],
        }
        
        # Deduction column patterns
        deduction_patterns = {
            'tax': ['tax', 'pph', 'pajak', 'income tax'],
            'bpjs': ['bpjs', 'bpjs kesehatan', 'bpjs ketenagakerjaan'],
            'jht': ['jht', 'jaminan hari tua'],
            'jp': ['jp', 'jaminan pensiun'],
            'jkk': ['jkk', 'jaminan kecelakan kerja'],
            'jkm': ['jkm', 'jaminan kematian'],
            'potongan': ['potongan', 'deduction', 'other deduction', 'lainnya'],
        }
        
        # Map headers to columns (A, B, C, ...)
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            header_lower = header.lower().strip()
            
            if not header_lower:
                continue
            
            # Check fixed columns
            for field, patterns in fixed_patterns.items():
                if any(pattern in header_lower for pattern in patterns):
                    mapping['fixed_columns'][field] = col_letter
            
            # Check earning columns
            for field, patterns in earning_patterns.items():
                if any(pattern in header_lower for pattern in patterns):
                    mapping['earnings'].append({'column': col_letter, 'key': field})
            
            # Check deduction columns
            for field, patterns in deduction_patterns.items():
                if any(pattern in header_lower for pattern in patterns):
                    mapping['deductions'].append({'column': col_letter, 'key': field})
        
        return mapping


# Create singleton instance untuk convenience
excel_parser = ExcelParser()


# Backward compatibility - keep MockExcelParser untuk testing
class MockExcelParser:
    """
    Mock parser untuk testing tanpa file Excel actual
    Gunakan untuk development / unit testing
    """
    def __init__(self):
        self.parser = ExcelParser()
    
    def parse_file(self, file_path: str, column_mapping: Dict = None) -> list:
        """
        Mock parser - returns dummy data
        Untuk actual parsing, gunakan ExcelParser class di atas
        """
        logger.info(f"📋 Mock parsing: {file_path}")
        
        # Return dummy data for testing
        return [
            {
                'row_number': 1,
                'employee_number': 'EMP-001',
                'full_name': 'John Doe',
                'earnings': {'salary': 5000000.0, 'allowance': 1000000.0},
                'deductions': {'tax': 500000.0, 'bpjs': 200000.0},
            },
            {
                'row_number': 2,
                'employee_number': 'EMP-002',
                'full_name': 'Jane Smith',
                'earnings': {'salary': 6000000.0, 'allowance': 1500000.0},
                'deductions': {'tax': 600000.0, 'bpjs': 250000.0},
            },
            {
                'row_number': 3,
                'employee_number': 'EMP-999',
                'full_name': 'Unknown User',
                'earnings': {'salary': 5000000.0},
                'deductions': {},
            },
        ]