from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from uuid import UUID
from datetime import datetime, date
from typing import Optional
import os
import shutil
import hashlib
from pydantic import BaseModel
from openpyxl import load_workbook

from app.config.database import get_db
from app.models.database import PayslipUploadSession, PayslipUploadRow
from app.repositories.company_repository import CompanyRepository
from app.repositories.upload_session_repository import UploadSessionRepository
from app.services.upload_service import UploadService
from app.utils.auth import get_current_user, CurrentUser, check_admin_access, UserRole
from app.utils.hashing import calculate_file_hash
from app.utils.excel_parser import ExcelParser
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/companies", tags=["uploads"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


class ColumnMappingRequest(BaseModel):
    fixed_columns: dict
    earnings: list
    deductions: list
    ignored_columns: Optional[list] = []
    start_row: Optional[int] = 2


class UploadSessionResponse(BaseModel):
    upload_session_id: str
    company_id: str
    file_name: str
    file_hash: str
    period_start: str
    period_end: str
    payment_date: str
    status: str
    created_at: str
    created_by: str


@router.post("/{company_id}/payslips/uploads")
async def upload_excel_file(
    company_id: UUID,
    period_start: str = Query(..., description="Period start date (YYYY-MM-DD)"),
    period_end: str = Query(..., description="Period end date (YYYY-MM-DD)"),
    payment_date: str = Query(..., description="Payment date (YYYY-MM-DD)"),
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Upload Excel file and create upload session
    
    POST /api/v1/companies/{company_id}/payslips/uploads
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    if current_user.company_id != company_id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied: Company mismatch")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "invalid_file_format",
                "message": "File must be Excel format (.xlsx or .xls)"
            }
        )
    
    try:
        period_start_date = datetime.strptime(period_start, "%Y-%m-%d").date()
        period_end_date = datetime.strptime(period_end, "%Y-%m-%d").date()
        payment_date_value = datetime.strptime(payment_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "invalid_date_format",
                "message": "Date must be in YYYY-MM-DD format"
            }
        )
    
    company_repo = CompanyRepository(db)
    if not company_repo.company_exists(company_id):
        raise HTTPException(status_code=404, detail="Company not found")
    
    file_content = await file.read()
    # Include period dates in hash to allow same file content with different periods
    file_hash_base = hashlib.sha256(file_content).hexdigest()
    file_hash = hashlib.sha256(f"{file_hash_base}_{period_start}_{period_end}_{payment_date}".encode()).hexdigest()
    
    # Check if there's already a session for the same period (not same file hash)
    existing_by_period = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.company_id == company_id,
        PayslipUploadSession.period_start == period_start_date,
        PayslipUploadSession.period_end == period_end_date,
        PayslipUploadSession.payment_date == payment_date_value
    ).first()
    
    if existing_by_period:
        # Update existing session instead of creating new
        existing_by_period.file_hash = file_hash
        existing_by_period.file_path = os.path.join(UPLOAD_DIR, f"{file_hash}_{file.filename}")
        existing_by_period.status = 'pending'
        db.commit()
        
        # Save new file
        file_path = existing_by_period.file_path
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        return {
            "success": True,
            "message": "Payslip data for this period has been updated",
            "upload_session": {
                "upload_session_id": str(existing_by_period.upload_session_id),
                "company_id": str(company_id),
                "file_name": file.filename,
                "file_hash": file_hash,
                "period_start": period_start,
                "period_end": period_end,
                "payment_date": payment_date,
                "status": existing_by_period.status,
                "created_at": existing_by_period.created_at.isoformat(),
                "created_by": str(current_user.user_id)
            }
        }
    
    file_path = os.path.join(UPLOAD_DIR, f"{file_hash}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    try:
        service = UploadService(db)
        upload_session = service.create_upload_session(
            company_id=company_id,
            file_path=file_path,
            file_hash=file_hash,
            period_start=period_start_date,
            period_end=period_end_date,
            payment_date=payment_date_value,
            created_by=str(current_user.user_id)
        )
        
        return {
            "success": True,
            "upload_session": {
                "upload_session_id": str(upload_session.upload_session_id),
                "company_id": str(company_id),
                "file_name": file.filename,
                "file_hash": file_hash,
                "period_start": period_start,
                "period_end": period_end,
                "payment_date": payment_date,
                "status": upload_session.status,
                "created_at": upload_session.created_at.isoformat(),
                "created_by": str(current_user.user_id)
            }
        }
    
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{company_id}/payslips/uploads/{upload_session_id}/auto-process")
def auto_process_upload(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Auto-detect columns and directly process upload
    
    POST /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/auto-process
    Auth: admin only
    """
    check_admin_access(current_user)
    
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied: Admin privileges required")
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    if upload_session.status not in ['pending', 'mapping_submitted']:
        raise HTTPException(
            status_code=409,
            detail=f"Session is already {upload_session.status}"
        )
    
    try:
        # Auto-detect columns
        parser = ExcelParser()
        column_mapping = parser.auto_detect_columns(upload_session.file_path)
        
        logger.info(f"Auto-detected column mapping: {column_mapping}")
        
        # Check if employee_number was detected
        if 'employee_number' not in column_mapping['fixed_columns']:
            return {
                "success": False,
                "error": "auto_detect_failed",
                "message": "Could not auto-detect 'employee_number' column. Please map columns manually.",
                "detected_mapping": column_mapping
            }
        
        # Check if any earnings were detected
        if not column_mapping['earnings']:
            # Try to detect a generic salary column
            wb = load_workbook(upload_session.file_path, data_only=True)
            ws = wb.active
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(cell).lower().strip() if cell else "" for cell in row]
            
            # Find column that might be salary (has 'gaji' or 'salary' in header)
            from openpyxl.utils import get_column_letter
            for i, header in enumerate(headers):
                if 'gaji' in header or 'salary' in header or 'total' in header:
                    col_letter = get_column_letter(i + 1)
                    column_mapping['earnings'].append({'column': col_letter, 'key': 'salary'})
                    break
        
        # Submit mapping (internal)
        upload_session.status = 'mapping_submitted'
        upload_session.column_mapping = column_mapping
        db.commit()
        
        # Process directly
        upload_service = UploadService(db)
        result = upload_service.process_upload(upload_session_id)
        
        return {
            "success": True,
            "message": "Payslip processed successfully",
            "result": {
                "total_rows": result.get('total_rows', 0),
                "successfully_inserted": result.get('successfully_inserted', 0),
                "failed": result.get('failed', 0),
            }
        }
    
    except Exception as e:
        logger.error(f"Error auto-processing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{company_id}/payslips/uploads/{upload_session_id}/preview")
def get_upload_preview(
    company_id: UUID,
    upload_session_id: UUID,
    rows: int = Query(5, ge=1, le=20),
    sheet: str = Query(None, description="Sheet name to preview"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get first rows of uploaded file for preview
    
    GET /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/preview
    Auth: admin only
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    try:
        parser = ExcelParser()
        headers, sample_rows = parser.get_preview(upload_session.file_path, max_rows=rows, sheet_name=sheet)
        all_sheets = parser.get_all_sheets(upload_session.file_path)
        
        return {
            "success": True,
            "preview": {
                "upload_session_id": str(upload_session_id),
                "total_columns": len(headers),
                "headers": headers,
                "sample_rows": [
                    {"row_number": i + 1, "data": row}
                    for i, row in enumerate(sample_rows)
                ],
                "all_sheets": all_sheets,
                "current_sheet": sheet or "default"
            }
        }
    
    except Exception as e:
        logger.error(f"Error reading preview: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{company_id}/payslips/uploads/{upload_session_id}/mapping")
def submit_column_mapping(
    company_id: UUID,
    upload_session_id: UUID,
    mapping: ColumnMappingRequest,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Submit column mapping from admin
    
    POST /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/mapping
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    if upload_session.status != 'pending':
        raise HTTPException(
            status_code=409,
            detail={
                "success": False,
                "error": "invalid_status",
                "message": f"Session is already {upload_session.status}"
            }
        )
    
    if not mapping.fixed_columns.get('employee_number'):
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "invalid_mapping",
                "message": "employee_number is required"
            }
        )
    
    if not mapping.earnings:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "invalid_mapping",
                "message": "At least one earning column must be specified"
            }
        )
    
    try:
        service = UploadService(db)
        
        column_mapping = {
            'start_row': mapping.start_row,
            'fixed_columns': mapping.fixed_columns,
            'earnings': mapping.earnings,
            'deductions': mapping.deductions,
            'ignored_columns': mapping.ignored_columns or []
        }
        
        updated_session = service.submit_column_mapping(upload_session_id, column_mapping)
        
        return {
            "success": True,
            "upload_session": {
                "upload_session_id": str(upload_session_id),
                "status": updated_session.status,
                "column_mapping": updated_session.column_mapping
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{company_id}/payslips/uploads/{upload_session_id}/process")
def start_processing(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Start processing upload session
    
    POST /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/process
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    if upload_session.status not in ['pending', 'mapped']:
        raise HTTPException(
            status_code=409,
            detail={
                "success": False,
                "error": "already_processing",
                "message": f"Upload is already {upload_session.status}",
                "job_status": upload_session.status
            }
        )
    
    try:
        service = UploadService(db)
        
        from uuid import uuid4
        job_id = str(uuid4())
        
        result = service.process_upload(upload_session_id)
        
        return {
            "success": True,
            "message": "Processing completed",
            "upload_session": {
                "upload_session_id": str(upload_session_id),
                "company_id": str(company_id),
                "status": "completed",
                "job_id": job_id,
                "job_status": "completed",
                "started_at": upload_session.created_at.isoformat(),
                "result": result
            }
        }
    
    except Exception as e:
        logger.error(f"Error processing upload: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{company_id}/payslips/uploads/{upload_session_id}/status")
def get_upload_status(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get real-time status of upload processing
    
    GET /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/status
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    total_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id
    ).count()
    
    successful_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'success'
    ).count()
    
    failed_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'failed'
    ).count()
    
    percentage = int((successful_rows + failed_rows) / total_rows * 100) if total_rows > 0 else 0
    
    return {
        "success": True,
        "upload_session": {
            "upload_session_id": str(upload_session_id),
            "status": upload_session.status,
            "progress": {
                "total_rows": total_rows,
                "processed_rows": successful_rows + failed_rows,
                "successful_inserts": successful_rows,
                "failed_rows": failed_rows,
                "percentage": percentage
            }
        }
    }


@router.get("/{company_id}/payslips/uploads/{upload_session_id}/result")
def get_upload_result(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get final results with error summary
    
    GET /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/result
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    failed_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'failed'
    ).all()
    
    errors = [
        {
            "row_number": row.row_number,
            "employee_number": row.employee_number,
            "error_type": "validation",
            "error_message": row.error_message
        }
        for row in failed_rows
    ]
    
    total_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id
    ).count()
    
    successful_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'success'
    ).count()
    
    status = "completed" if len(errors) == 0 else "completed_with_errors"
    if upload_session.status == "failed":
        status = "failed"
    
    return {
        "success": True,
        "result": {
            "upload_session_id": str(upload_session_id),
            "status": status,
            "summary": {
                "total_rows": total_rows,
                "successfully_inserted": successful_rows,
                "failed": len(errors)
            },
            "errors": errors,
            "completed_at": upload_session.updated_at.isoformat() if upload_session.status in ['completed', 'completed_with_errors'] else None
        }
    }


@router.get("/{company_id}/payslips/uploads/{upload_session_id}/failed-rows")
def get_failed_rows(
    company_id: UUID,
    upload_session_id: UUID,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get list of failed rows from an upload session
    
    GET /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/failed-rows
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    total = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'failed'
    ).count()
    
    failed_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'failed'
    ).offset((page - 1) * per_page).limit(per_page).all()
    
    return {
        "success": True,
        "data": [
            {
                "row_number": row.row_number,
                "employee_number": row.employee_number,
                "full_name": row.raw_data.get('full_name') if row.raw_data else 'Unknown',
                "error_message": row.error_message,
                "raw_data": row.raw_data
            }
            for row in failed_rows
        ],
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }


@router.get("/{company_id}/payslips/uploads/history")
def get_upload_history(
    company_id: UUID,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get all uploads for a company
    
    GET /api/v1/companies/{company_id}/payslips/uploads/history
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    if current_user.company_id != company_id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied")
    
    total = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.company_id == company_id
    ).count()
    
    sessions = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.company_id == company_id
    ).order_by(
        PayslipUploadSession.created_at.desc()
    ).offset((page - 1) * per_page).limit(per_page).all()
    
    data = []
    for session in sessions:
        total_rows = db.query(PayslipUploadRow).filter(
            PayslipUploadRow.upload_session_id == session.id
        ).count()
        
        successful_rows = db.query(PayslipUploadRow).filter(
            PayslipUploadRow.upload_session_id == session.id,
            PayslipUploadRow.status == 'success'
        ).count()
        
        failed_rows = db.query(PayslipUploadRow).filter(
            PayslipUploadRow.upload_session_id == session.id,
            PayslipUploadRow.status == 'failed'
        ).count()
        
        data.append({
            "upload_session_id": str(session.upload_session_id),
            "parent_upload_session_id": str(session.parent_upload_session_id) if session.parent_upload_session_id else None,
            "period_start": session.period_start.isoformat(),
            "period_end": session.period_end.isoformat(),
            "file_name": os.path.basename(session.file_path),
            "status": session.status,
            "summary": {
                "total_rows": total_rows,
                "successfully_inserted": successful_rows,
                "failed": failed_rows
            },
            "created_at": session.created_at.isoformat(),
            "completed_at": session.updated_at.isoformat() if session.status in ['completed', 'completed_with_errors'] else None,
            "created_by": session.created_by
        })
    
    return {
        "success": True,
        "data": data,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total_items": total,
            "total_pages": (total + per_page - 1) // per_page
        }
    }


@router.get("/{company_id}/payslips/uploads/{upload_session_id}/failed-rows/download-csv")
def download_failed_rows_csv(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Download failed rows as CSV
    
    GET /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/failed-rows/download-csv
    
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    if current_user.company_id != company_id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied: Company mismatch")
    
    upload_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not upload_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    failed_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == upload_session.id,
        PayslipUploadRow.status == 'failed'
    ).all()
    
    if not failed_rows:
        raise HTTPException(status_code=404, detail="No failed rows to download")
    
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['row_number', 'employee_number', 'full_name', 'error_message', 'raw_data'])
    
    for row in failed_rows:
        full_name = row.raw_data.get('full_name', 'Unknown') if row.raw_data else 'Unknown'
        raw_data_str = ''
        if row.raw_data:
            raw_data_str = str(row.raw_data).replace('"', '""')
        
        writer.writerow([
            row.row_number,
            row.employee_number,
            full_name,
            row.error_message or '',
            raw_data_str
        ])
    
    csv_content = output.getvalue()
    
    from fastapi.responses import StreamingResponse
    
    filename = f"failed_rows_{upload_session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([csv_content]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.post("/{company_id}/payslips/uploads/{upload_session_id}/reprocess")
def reprocess_failed_rows(
    company_id: UUID,
    upload_session_id: UUID,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Reprocess failed rows from an upload session
    
    POST /api/v1/companies/{company_id}/payslips/uploads/{upload_session_id}/reprocess
    
    Creates a child session and processes only the failed rows.
    
    Auth: admin, client_admin
    """
    check_admin_access(current_user)
    
    if current_user.company_id != company_id and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied: Company mismatch")
    
    parent_session = db.query(PayslipUploadSession).filter(
        PayslipUploadSession.upload_session_id == upload_session_id,
        PayslipUploadSession.company_id == company_id
    ).first()
    
    if not parent_session:
        raise HTTPException(status_code=404, detail="Upload session not found")
    
    if parent_session.status not in ['completed', 'completed_with_errors']:
        raise HTTPException(
            status_code=409,
            detail={
                "success": False,
                "error": "invalid_status",
                "message": f"Cannot reprocess session with status: {parent_session.status}"
            }
        )
    
    failed_rows = db.query(PayslipUploadRow).filter(
        PayslipUploadRow.upload_session_id == parent_session.id,
        PayslipUploadRow.status == 'failed'
    ).all()
    
    if not failed_rows:
        raise HTTPException(
            status_code=404,
            detail={
                "success": False,
                "error": "no_failed_rows",
                "message": "No failed rows to reprocess"
            }
        )
    
    try:
        from uuid import uuid4
        from app.services.upload_service import UploadService
        
        child_session_id = uuid4()
        
        child_session = PayslipUploadSession(
            id=child_session_id,
            upload_session_id=child_session_id,
            company_id=parent_session.company_id,
            file_path=parent_session.file_path,
            file_hash=parent_session.file_hash,
            parent_upload_session_id=parent_session.id,
            period_start=parent_session.period_start,
            period_end=parent_session.period_end,
            payment_date=parent_session.payment_date,
            status='pending',
            column_mapping=parent_session.column_mapping,
            created_by=str(current_user.user_id)
        )
        
        db.add(child_session)
        db.commit()
        
        upload_service = UploadService(db)
        
        column_mapping = parent_session.column_mapping or {}
        
        reprocessed_data = []
        for row in failed_rows:
            if row.raw_data:
                reprocessed_data.append(row.raw_data)
        
        result = upload_service.process_reprocess(
            child_session_id=child_session_id,
            column_mapping=column_mapping,
            data_rows=reprocessed_data
        )
        
        return {
            "success": True,
            "message": "Reprocessing started",
            "child_upload_session": {
                "upload_session_id": str(child_session_id),
                "parent_upload_session_id": str(upload_session_id),
                "status": "inserting",
                "rows_to_process": len(reprocessed_data),
                "started_at": child_session.created_at.isoformat() if child_session.created_at else None
            }
        }
    
    except Exception as e:
        logger.error(f"Error reprocessing failed rows: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
